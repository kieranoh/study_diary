from openpyxl import load_workbook
from openpyxl import Workbook

# Chapter 6 (search)

wb = load_workbook("sample.xlsx")
ws = wb.active

for row in ws.iter_rows(min_row=2):
    if int(row[1].value) > 80:
        print(row[0].value, "is good at eng")

for row in ws.iter_rows(max_row = 1):
    for cell in row:
        if cell.value == "English":
            cell.value = "Computer"

wb.save("sample_modified.xlsx")

'''
Summary 

'''

# Chapter 7 (delete)

wb = load_workbook("sample.xlsx")
ws = wb.active

# ws.delete_rows(8) # delete row number 8
# ws.delete_rows(8,3) # delete 3 rows from row number 8
# wb.save("sample_delete_row.xlsx")

# ws.delete_cols(2) # delete column B
ws.delete_cols(2,2) # delete 2 columns from clomun B

wb.save("sample_delete_col.xlsx")

'''
Summary 

ws.delete_rows(row_number, how many)  -> delete the rows from row_number
ws.delete_cols(column_number, how many) -> delete the columns from column_number 
'''

# Chapter 9 (move)

wb = load_workbook("sample.xlsx")
ws = wb.active

ws.move_range("B1:C11",rows=0,cols=1) # row = -1  -> move up

ws["B1"] = "Korean"

wb.save("sample_korean.xlsx")

'''
Summary 

ws.move_range("A1:C3",rows = -2, cols = 2)  -> move the datas in A1:C3 row - 2( == up) col 2 ( == right) 
'''

# Chapter 10 (chart)

from openpyxl.chart import BarChart, Reference, LineChart
wb = load_workbook("sample.xlsx")
ws = wb.active

# bar_value = Reference(ws,min_row = 2, max_row = 11, min_col = 2, max_col = 3)
# bar_chart = BarChart() # bar chart, line chart, pie chart....
# bar_chart.add_data(bar_value)
# ws.add_chart(bar_chart, "E1")

line_value = Reference(ws,min_row = 1, max_row = 11, min_col = 2, max_col = 3)
line_chart = LineChart()
line_chart.add_data(line_value, titles_from_data = True)
line_chart.title = "Score chart"
line_chart.style = 20 # style that are saved already
line_chart.y_axis.title = "score" 
line_chart.x_axis.title = "number"

ws.add_chart(line_chart,"E1")
wb.save("sample_chart_line.xlsx")

'''
Summary 

value = Reference(ws,min_row = 2, max_row = 11, min_col = 2, max_col = 3) 
-> read the data
Barchart() -> make bar chart
chart.add_data(value, titles_from_data) -> input the value in chart ( read the title)
chart.title = "chart name"  -> set chart name
chart.style = "10" -> set chart style 
chart.x_axis.title = "x"  -> set x axis title
chart.y_axis.title = "y"  -> set y axis title
ws.add_chart(chart, position) -> add chart in the sheet

'''

# chapter 11 (cell style)

from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
wb = load_workbook("sample.xlsx")
ws = wb.active

a1 = ws["A1"]
b1 = ws["B1"]
c1 = ws["C1"]

ws.column_dimensions["A"].width = 8 # set column A width
ws.row_dimensions[1].height = 50

a1.font = Font(color="FF0000", italic= True, bold= True) # set red color, italic and bold
b1.font = Font(color="CC33FF", name="Arial", strike=True) # set fint arial and strike
c1.font = Font(color="0000FF", size = 20, underline="single") # set size and make one underline

# set border
thin_border = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
a1.border = thin_border
b1.border = thin_border
c1.border = thin_border

for row in ws.rows:
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center") # center, left,right, top,bottom.....
        if cell.column == 1: # column A
            continue

        if isinstance(cell.value, int) and cell.value >= 90: # if cell.value is int and bigger than 90
            cell.fill = PatternFill(fgColor="00ff00", fill_type="solid")
            cell.font = Font(color="ff0000")


ws.freeze_panes = "B2"
wb.save("sample_style.xlsx")

'''
Summary 

ws.column_dimensions["A"].width = 5  -> set column A width to 5

a1 = ws["A1"]
a1.font = Font(color="00FF00", name = "font name", italic = True
                bold = True, strike = True, size = 20, underline="single)
-> set font

border = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
-> set border   all side = thin
a1.border = border  

cell.alignment = Alignment(horizontal="center", vertical="center") # center, left,right, top,bottom.....
-> set position to center

isinstance(cell.value, int)  -> check value is int or not

cell.fill = PatternFill(fgColor="00ff00", fill_type="solid") -> fill the color to cell

ws.freeze_panes = "B2"  -> freeze the cell for easy to see the file
'''

# Chapter 12 (formula)
import datetime

wb = Workbook()
ws= wb.active

ws["A1"] = datetime.datetime.today()
ws["A2"] = "=SUM(1,2,3)"
ws["A3"] = "=AVERAGE(1,2,3)"


ws["A4"] = 10
ws["A5"] = 20
ws["A6"] = "=SUM(A4:A5)"

wb.save("sample_formula.xlsx")

'''
Summary

datetime.datetime.today()  -> get time

ws["A2"] = "=SUM(1,2,3)"
ws["A3"] = "=AVERAGE(1,2,3)" 
-> input the formula that use in xlsx
'''

# Chapter 13 (read data only)

wb = load_workbook("sample_formula.xlsx")
ws = wb.active

# read the formula
for row in ws.values:
    for cell in row:
        print(cell)

wb = load_workbook("sample_formula.xlsx", data_only=True)
ws = wb.active

# read the data (not evluated data will show as None)
for row in ws.values:
    for cell in row:
        print(cell)

'''
Summary

wb = load_workbook("sample_formula.xlsx", data_only=True)
-> set data_only = True for read the data (result of the formula not the formula)
'''

# Chapter 14 (merge)

wb =Workbook()
ws = wb.active

ws.merge_cells("B2:D2") # merge the cell
ws["B2"] = "Merged Cell"

wb.save("sample_merge.xlsx")

'''
Summary

ws.merge_cells(area) -> merge the cell
'''

# Chapter 15 (unmerge)

wb = load_workbook("sample_merge.xlsx")
ws = wb.active

ws.unmerge_cells("B2:D2")

wb.save("sample_unmerge.xlsx")

'''
Summary

ws.unmerge_cells(area) -> unmerge the cell
'''

# Chapter 16 (image)

from openpyxl.drawing.image import Image
wb = Workbook()
ws = wb.active

img = Image("img.png")
ws.add_image(img,"C3")

wb.save("sample_image.xlsx")

'''
Summary

img = Image("image name") -> take the image
ws.add_image(img,position) -> add the image to the sheet
'''

# QUIZE
'''
each grade are saved at score.xlsx
[number, attendence point, quize 1, quize 2, mid term, final, report]
change the data to 10 in quize 2
add the column H for sum (use the formula)
add the coolumn I for grade
if the attendence point is smaller than 5  -> grade =5
'''
from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string
wb = load_workbook("scores.xlsx")
ws = wb.active

# set quize 2 point to 10

for col in ws.iter_cols():
    if col[0].value == "퀴즈2":
        for cell in col[1:]:
            cell.value = 10

# set column H, I
# = SUM(B1:G1)
ws["H1"] = "총점"
ws["I1"] = "성적"
for row in ws.iter_rows(min_row=2,min_col=2):
    sum = 0
    x = coordinate_from_string(row[0].coordinate)[1]
    ws.cell(column=8, row=x, value="=SUM(B{}:G{})".format(x,x))
    if int(row[0].value) < 5:
        ws.cell(column=9, row=x, value="F")
        continue

    for cell in row[:-2]:
        sum += int(cell.value)
    if sum >= 90:
        ws.cell(column=9, row=x, value="A")
    elif sum >= 80:
        ws.cell(column=9, row=x, value="B")
    elif sum >= 70:
        ws.cell(column=9, row=x, value="C")
    else:
        ws.cell(column=9, row=x, value="D")
    print(x,sum)

wb.save("score.xlsx")
