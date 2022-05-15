from openpyxl import Workbook
# Chapter 1
wb = Workbook() # create new workbook
ws = wb.active # take active sheet
ws.title = "NadoSheet"
wb.save("sample.xlsx")
wb.close()

'''
Summary 

Workbook() -> create file
wb.active -> get the active sheet
ws.title -> set the title name
''' 

# Chapter 2
wb = Workbook()
ws = wb.create_sheet()
ws.title = "MySheet"
# change the sheet tab color using RGB (google RGB seach the color)
ws.sheet_properties.tabColor = "ff66ff"  
ws1 = wb.create_sheet("yoursheet")
ws2 = wb.create_sheet("new",2) # set index to set location

new_ws = wb["new"] # using dict to select the sheet
print(wb.sheetnames) # get all sheet name

new_ws["A1"] = "Test" # input the data to A1
target = wb.copy_worksheet(new_ws) # copy the sheet and make the sheet that have the same data
target.title = "Copied Sheet"

print(wb.sheetnames)
wb.save("sample.xlsx")

'''
Summary 

ws.sheet_properties.tabcolor = "RGB" -> set the sheet tab color
ws.create_sheet("name",index) -> create the new sheet if there is not index value it will save at last
all sheet name will save as dicto (we can call the sheet using dict)
ws["A1"] = "data" -> save the data at A1
ws2.copy_worksheet(ws) ->copy the ws sheet
wb.sheetnames -> show the all sheet name
''' 
# Cahpter 3
from random import *
wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

# input the data
ws["A1"] = 1
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

print(ws["A1"]) # output ->   <Cell 'NadoSheet'.A1>
print(ws["A1"].value) # use .value to get the data
print(ws["A10"].value) # output ->  None

# row = 1 , 2 , 3
# column = A , B , C
print(ws.cell(row=1, column=1).value) # == ws.["A1"]

c = ws.cell(column=3, row=1, value=10) # == ws.["C1"] = 10
print(c.value)
idx = 1
for x in range(1,11):
    for y in range(1,11):
        # ws.cell(row=x,column=y,value=randint(0,100))
        ws.cell(row=x,column=y,value=idx)
        idx += 1
wb.save("sample.xlsx")

'''
Summary 

ws["A1"] -> <Cell 'Sheetname'.A1>  (use .value to get the cewll value)
ws.cell(row=num, column=num) == ws["~~"] (column = 1, row =1    ==    A1)
''' 
# Chapter 4
from openpyxl import *
# read the xlsx file that already saved
wb = load_workbook("sample.xlsx")
ws = wb.active
for x in range(1,11):
    for y in range(1,11):
        print(ws.cell(row=x,column=y).value, end=" ")
    print()

# use max_row and max_column when we do not know the file range
for x in range(1,ws.max_row+1):
    for y in range(1,ws.max_column+1):
        print(ws.cell(row=x, column=y).value, end=" ")
    print()

'''
Summary 

load_workbook("filename) -> load the file
''' 
# Chapter 5
from random import *
from openpyxl.utils.cell import coordinate_from_string

wb = Workbook()
ws = wb.active
# append the value with row
ws.append(["number","English","Math"])
for i in range(1,11):
    ws.append([i,randint(0,100),randint(0,100)])

col_B = ws["B"] # get the B column
print(col_B)
for cell in col_B:
    print(cell.value)

col_range = ws["B:C"] # get column B to C

for clos in col_range:
    # get the column
    for cell in clos:
        print(cell.value)

row_title = ws[1] # get the row 1
for cell in row_title:
    print(cell.value)

row_range = ws[2:6] #get the row 2 to 6 ( 6 also )

for rows in row_range:
    for cell in rows:
        print(cell.value, end = " ")
    print()

new_row_range = ws[2:ws.max_row]
for rows in new_row_range:
    for cell in rows:
        print(cell.value, end = " ")
        print(cell.coordinate, end=" ")
        xy = coordinate_from_string(cell.coordinate) # get the position by tuple
        print(xy, end=" ")
        print(xy[0], end="")
        print(xy[1],end="   ")
    print()
print(tuple(ws.rows))
print(tuple(ws.columns))

for row in tuple(ws.rows):
    print(row[1].value)

for row in ws.iter_rows():
    print(row[1].value)

for row in ws.iter_rows(min_row=1, max_row=5, min_col=2,max_col=3): # set the area
    print(row[0].value, row[1].value)

wb.save("sample.xlsx")

'''
Summary

ws.append(~~) -> append the data using row
ws["B"] -> get all data in column B (using "for" to read each data)
ws["B:C"] -> get all data in column B to C (read as column)
ws[1] -> get all data in row 1
cell.coordinate -> get the cell name    ex) A1, B2......
coordinate_from_string(cell.coordinate)  -> get the cell name as tuple   ex)  ('A',1).....
tuple(ws.rows)  -> get each row data as tuple  (use "for" to get each data)
tuple(ws.columns)  -> get each column data as tuple
ws.iter_row == tuple(ws.rows), ws.iter_column == typle(ws.columns)
ws.iter_rows(min_row=~, max_row=~, min_col=~, max_col=~) -> set the area

'''